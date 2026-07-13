"""Generate ../corporate-data.js directly from the rating workbook so the web app's
data layer is a faithful copy of the spreadsheet's "Look Up Tables".

Run from anywhere:  py "Corporate Rating Engine/tools/gen_corporate_data.py"
The workbook is expected beside the app folder; if absent (e.g. running inside a
git worktree where the .xlsx is untracked) it falls back to the main-repo copy.
"""
import json
from pathlib import Path
import openpyxl

HERE = Path(__file__).resolve().parent            # .../Corporate Rating Engine/tools
APP = HERE.parent                                 # .../Corporate Rating Engine
XLSX_NAME = "Phishield Corporate Rarting Engine with Sec Questions_13_03_2026.xlsx"
DEFAULT_XLSX = APP / XLSX_NAME
FALLBACK_XLSX = Path(r"C:\Users\sarel\Desktop\Sarel\SML Consulting\PSQ\Corporate Rating Engine") / XLSX_NAME
PATH = DEFAULT_XLSX if DEFAULT_XLSX.exists() else FALLBACK_XLSX
OUT = APP / "corporate-data.js"

wb = openpyxl.load_workbook(str(PATH), data_only=True)
lut = wb["Look Up Tables"]


def col(letter, r1, r2):
    return [lut[f"{letter}{r}"].value for r in range(r1, r2 + 1)]


# ---- Base premium constants (cover -> constant) A3:A11 / B3:B11 ----
covers = col("A", 3, 11)
consts = col("B", 3, 11)
base_premium = [{"cover": c, "constant": (None if isinstance(k, str) else k),
                 "raw": (k if isinstance(k, str) else None)} for c, k in zip(covers, consts)]
cover_options = [c for c, k in zip(covers, consts) if not isinstance(k, str)]

# ---- Sub-industry factor table N3:N88 ----
industries = []
for r in range(3, 89):
    sub = lut[f"N{r}"].value
    if sub is None:
        continue
    industries.append({
        "main": lut[f"M{r}"].value,
        "sub": sub,
        "breachUSD": lut[f"O{r}"].value,
        "breachZAR": lut[f"P{r}"].value,
        "industryFac": lut[f"S{r}"].value,
        "biFac": lut[f"U{r}"].value,
        "row": r,
    })
depository_sub = lut["N57"].value

# ---- Benefits AD4:AD12 / AG4:AG12 ----
benefits = [{"name": n, "contribution": g} for n, g in zip(col("AD", 4, 12), col("AG", 4, 12))]

# ---- Maturity bands AP3:AT8 ----
maturity = [{"label": lut[f"AP{r}"].value, "multiplier": lut[f"AQ{r}"].value,
             "description": lut[f"AR{r}"].value, "gte": lut[f"AS{r}"].value,
             "lt": lut[f"AT{r}"].value} for r in range(3, 9)]

# ---- Funds Protect (both tables straight from the workbook) ----
# The actual corporate FP contributions run from R500k upward (Summary D11 dropdown -> A74:J75).
fp_standard = {"amounts": [lut.cell(row=69, column=c).value for c in range(1, 12)],
               "costs":   [lut.cell(row=70, column=c).value for c in range(1, 12)]}
fp_adjustable = {"amounts": [lut.cell(row=74, column=c).value for c in range(1, 11)],
                 "costs":   [lut.cell(row=75, column=c).value for c in range(1, 11)]}

excess_options = [v for v in col("A", 15, 65) if v is not None]
mdr = [{"label": lut[f"A{r}"].value, "discount": lut[f"B{r}"].value} for r in range(89, 94)]
depository_bands = [{"modifier": lut[f"A{r}"].value, "gte": lut[f"B{r}"].value,
                     "lt": lut[f"C{r}"].value} for r in range(79, 87)]

J3, J4 = lut["J3"].value, lut["J4"].value
constants = {
    "EXPONENT_A": -0.03035, "EXPONENT_B": 1.462732,
    "C15_DIVISOR": 1.155, "DEFAULT_VAT": 0.15, "RANSOM_DIV": 100,
    "YEARLY_MARKET_ADJ": -0.20, "EXCESS_POWER": 1.1, "EXCESS_HALF_COVER": 0.5,
    "CLAIMS_PORTION": 0.3, "RISK_MGMT_FEE": 0.06,
    "SA_BREACH_FACTOR": J4 / J3, "GLOBAL_BREACH_ZAR": J3, "SA_BREACH_ZAR": J4,
    "SME_CORP_RATIO": 0.6173479179, "TURNOVER_MIN": 250000000,
}
vat_options = [v for v in col("E", 2, 12) if v is not None]
benefit_sublimit_ratios = [lut.cell(row=15, column=c).value for c in range(31, 41)]

data = {
    "META": {"product": "Cyber Protect — Corporate (Risk Rated)",
             "administrator": "Phishield UMA (Pty) Ltd", "source_workbook": XLSX_NAME},
    "CONSTANTS": constants, "BASE_PREMIUM": base_premium, "COVER_OPTIONS": cover_options,
    "INDUSTRIES": industries, "DEPOSITORY_SUB": depository_sub, "DEPOSITORY_BANDS": depository_bands,
    "BENEFITS": benefits, "BENEFIT_SUBLIMIT_RATIOS": benefit_sublimit_ratios, "MATURITY_BANDS": maturity,
    "FP_STANDARD": fp_standard, "FP_ADJUSTABLE": fp_adjustable, "EXCESS_OPTIONS": excess_options,
    "MDR_OPTIONS": mdr, "VAT_OPTIONS": vat_options,
}

js = (
    "/* corporate-data.js — Corporate Rating Engine data layer.\n"
    " * AUTO-GENERATED from the workbook 'Look Up Tables' by tools/gen_corporate_data.py.\n"
    " * Do not edit by hand; re-run the generator if the spreadsheet changes.\n"
    " */\n"
    "const CORP_DATA = " + json.dumps(data, indent=2, ensure_ascii=False) + ";\n\n"
    "if (typeof window !== 'undefined') { window.CORP_DATA = CORP_DATA; }\n"
    "if (typeof module !== 'undefined' && module.exports) { module.exports = CORP_DATA; }\n"
)
OUT.write_text(js, encoding="utf-8")
print(f"Wrote {OUT}  (source: {PATH.name})")
print(f"  cover_options={cover_options}")
print(f"  industries={len(industries)}  benefits={len(benefits)}  maturity={len(maturity)}")
