"""Spreadsheet-parity check: re-derive the premium chain in Python directly from the
workbook's inputs + Look Up Tables, and compare each step to the workbook's own cached
values. If this passes, the engine's *formula structure* still matches the spreadsheet
(only the data changed). Run after any spreadsheet update, before regenerating the data layer.

    py "Corporate Rating Engine/tools/reproduce_engine.py"
"""
import math
from pathlib import Path
import openpyxl

HERE = Path(__file__).resolve().parent
APP = HERE.parent
XLSX_NAME = "Phishield Corporate Rarting Engine with Sec Questions_13_03_2026.xlsx"
PATH = APP / XLSX_NAME
if not PATH.exists():
    PATH = Path(r"C:\Users\sarel\Desktop\Sarel\SML Consulting\PSQ\Corporate Rating Engine") / XLSX_NAME

wb = openpyxl.load_workbook(str(PATH), data_only=True)
lut, prem, summ = wb["Look Up Tables"], wb["Premium Calculation"], wb["Summary & Input Flieds"]


def col(ws, L, r1, r2):
    return [ws[f"{L}{r}"].value for r in range(r1, r2 + 1)]


def xlookup(key, keys, vals, d=None):
    for k, v in zip(keys, vals):
        if k == key:
            return v
    return d


turnover, cover, subind = summ["C4"].value, summ["C5"].value, summ["C7"].value
override, vat = summ["C15"].value, summ["C31"].value
fp_adj, excess, mdr = summ["D11"].value, summ["D18"].value, summ["D34"].value

A3_10, B3_10 = col(lut, "A", 3, 10), col(lut, "B", 3, 10)
N, P, S, U = col(lut, "N", 3, 88), col(lut, "P", 3, 88), col(lut, "S", 3, 88), col(lut, "U", 3, 88)
AD, AG = col(lut, "AD", 4, 12), col(lut, "AG", 4, 12)
AP, AQ = col(lut, "AP", 3, 8), col(lut, "AQ", 3, 8)
AS, AT, AQ7 = col(lut, "AS", 3, 7), col(lut, "AT", 3, 7), col(lut, "AQ", 3, 7)
J3, J4 = lut["J3"].value, lut["J4"].value

passed = 0
failed = 0


def check(label, got, want):
    global passed, failed
    ok = want is not None and abs((got or 0) - (want or 0)) < max(1.0, abs(want or 0) * 1e-4)
    passed += ok; failed += (not ok)
    print(f"  [{'OK ' if ok else 'XX '}] {label:42s} {got:>18,.4f}  vs cached {want:>18,.4f}")


C6 = xlookup(cover, A3_10, B3_10) * (turnover / cover) ** (-0.03035 * math.log(turnover) + 1.462732)
check("C6 base premium", C6, prem["C6"].value)
C8, C9 = xlookup(subind, N, S), xlookup(subind, N, U)
C13 = xlookup(override, AP, AQ)
mat = C13 if override != AP[5] else None
C15 = C6 * (1 + C8) * (1 + C9) * mat / 1.155 * (1 + vat)
check("C15 adjusted premium", C15, prem["C15"].value)
D25 = xlookup("Cyber Extortion Costs", AD, AG)
C16 = C15 / 100 + D25 + C15
check("C16 ransom-incl", C16, prem["C16"].value)

names = [("Business Interruption Loss", "C21"), ("Multimedia Liability Claims", "C22"),
         ("Regulatory Expenses and Penalties", "C23"), ("Third Party Claims", "C24"),
         ("Emergency Response Costs", "C25"), ("Data Restoration Costs", "C26"),
         ("Cyber Extortion Costs", "C27"), ("PCI Fines and Penalties", "C28"),
         ("Computer Crime", "C29")]
H = 0.0
for nm, flag in names:
    if summ[flag].value:
        H += (xlookup(subind, N, U) if nm == "Business Interruption Loss" else (xlookup(nm, AD, AG) or 0))
C31 = C16 * H * (1 + prem["D31"].value)
check("C31 yearly market adj", C31, prem["C31"].value)

C42 = J4 / J3
C45 = xlookup(subind, N, P) * C42
C38 = (excess / C45) if C45 < cover else (excess / cover) ** 1.1
C48 = C31 * (1 - C38)
check("C48 post-excess", C48, prem["C48"].value)

C35 = xlookup(fp_adj, [lut.cell(74, c).value for c in range(1, 11)], [lut.cell(75, c).value for c in range(1, 11)])
C53 = (C48 + C35) / (1 - prem["C50"].value)
check("C53 base premium", C53, prem["C53"].value)
E56 = xlookup(mdr, col(lut, "A", 89, 93), col(lut, "B", 89, 93))
C57 = C53 - C53 * (E56 or 0)
check("C57 FINAL PREMIUM", C57, prem["C57"].value)

print(f"\n{passed} passed, {failed} failed — {'PARITY OK' if failed == 0 else 'PARITY BROKEN'}")
raise SystemExit(1 if failed else 0)
