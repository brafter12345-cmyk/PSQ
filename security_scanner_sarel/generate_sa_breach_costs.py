from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Sheet 1: SA Industry Breach Costs ──
ws = wb.active
ws.title = "SA Breach Costs by Industry"

NAVY = "0F2744"
BLUE = "1D4ED8"
WHITE = "FFFFFF"
GREEN_BG = "DCFCE7"
AMBER_BG = "FEF9C3"
RED_BG = "FEE2E2"
GREY_BG = "F1F5F9"
BLACK = "0F172A"

hdr_font = Font(name="Arial", bold=True, color=WHITE, size=11)
hdr_fill = PatternFill("solid", fgColor=NAVY)
hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
body_font = Font(name="Arial", size=10, color=BLACK)
bold_font = Font(name="Arial", size=10, color=BLACK, bold=True)
blue_font = Font(name="Arial", size=10, color=BLUE)
curr_fmt = '#,##0'
curr_fmt_m = 'R#,##0.0,,"M"'
pct_fmt = '0.00"x"'
thin = Side(style="thin", color="E2E8F0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Title block
ws.merge_cells("A1:H1")
ws["A1"] = "South Africa Data Breach Cost by Industry — Derived from IBM 2025 Report"
ws["A1"].font = Font(name="Arial", bold=True, size=14, color=NAVY)
ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 30

ws.merge_cells("A2:H2")
ws["A2"] = "Base: R44.1M avg breach cost | 23,445 avg records | R1,881/record | Exchange rate: R16.50/USD"
ws["A2"].font = Font(name="Arial", size=9, color="64748B")
ws.row_dimensions[2].height = 20

ws.merge_cells("A3:H3")
ws["A3"] = "Sources: IBM/Ponemon Cost of a Data Breach Report 2025, TechCentral SA, iAfrica"
ws["A3"].font = Font(name="Arial", italic=True, size=8, color="94A3B8")

# Headers row 5
headers = [
    "Industry",
    "Global Avg\nBreach Cost (USD)",
    "SA Derived\nBreach Cost (ZAR)",
    "SA Cost\nPer Record (ZAR)",
    "Industry\nMultiplier",
    "YoY\nTrend",
    "Data\nSource",
    "Notes",
]
col_widths = [28, 18, 18, 18, 14, 10, 14, 40]

for c, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=5, column=c, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = hdr_align
    cell.border = border
    ws.column_dimensions[get_column_letter(c)].width = w

ws.row_dimensions[5].height = 36

# Data — (industry, global_usd_m, sa_zar_m or None=derived, trend, source, notes)
# sa_zar_m=None means derive from global; actual IBM SA values provided directly
industries = [
    ("Healthcare",              7.42, None,  "Down",  "IBM Global",    "Highest globally for 14th consecutive year; 279-day containment"),
    ("Financial Services",      5.56, 70.2,  "Down",  "IBM SA Direct", "Highest in SA; POPIA enforcement driving costs"),
    ("Industrial / Manufacturing", 5.00, None, "Down", "IBM Global",   "High costs especially in APAC region"),
    ("Energy",                  4.83, None,  "Down",  "IBM Global",    "Critical infrastructure; regulatory penalties"),
    ("Technology",              4.79, None,  "Down",  "IBM Global",    "Supply chain exploits driving costs"),
    ("Pharmaceuticals",         4.61, None,  "Down",  "IBM Global",    "IP theft and regulatory compliance costs"),
    ("Services",                4.56, 56.8,  "Down",  "IBM SA Direct", "SA-specific: disproportionately higher than global ratio"),
    ("Entertainment",           4.43, None,  "Up",    "IBM Global",    "Bucked downward trend; costs increased YoY"),
    ("Media",                   4.22, None,  "Up",    "IBM Global",    "Bucked downward trend; costs increased YoY"),
    ("Hospitality",             4.03, 57.5,  "Up",    "IBM SA Direct", "SA-specific: significantly higher than global ratio"),
    ("Transportation",          3.98, None,  "Down",  "IBM Global",    "Logistics and supply chain data exposure"),
    ("Education",               3.80, None,  "Up",    "IBM Global",    "Increasing attacks on institutions; costs rising"),
    ("Research",                3.79, None,  "Up",    "IBM Global",    "IP and research data theft"),
    ("Communications",          3.75, None,  "Down",  "IBM Global",    "Telecom and ISP breach costs"),
    ("Consumer",                3.72, None,  "Down",  "IBM Global",    "PII and payment data exposure"),
    ("Retail",                  3.54, None,  "Up",    "IBM Global",    "Ransomware up 58% Q1-Q2 2025"),
    ("Public Sector",           2.86, None,  "Up",    "IBM Global",    "Lowest cost but increasing trend; government data"),
]

SA_AVG_ZAR_M = 44.1
GLOBAL_AVG_USD_M = 4.44
SA_RECORDS = 23445
USD_ZAR = 16.50
# SA adjustment ratio: how SA costs relate to global when converted
SA_RATIO = SA_AVG_ZAR_M / (GLOBAL_AVG_USD_M * USD_ZAR)  # ~0.602

# Reference cells for formulas
ws["J1"] = "SA Avg Breach (ZAR M)"
ws["K1"] = SA_AVG_ZAR_M
ws["J2"] = "Global Avg (USD M)"
ws["K2"] = GLOBAL_AVG_USD_M
ws["J3"] = "SA Avg Records"
ws["K3"] = SA_RECORDS
ws["J4"] = "USD/ZAR"
ws["K4"] = USD_ZAR
ws["J5"] = "SA Ratio"
ws["K5"] = f"=$K$1/($K$2*$K$4)"
for r in range(1, 6):
    ws.cell(row=r, column=10).font = Font(name="Arial", size=8, color="94A3B8")
    ws.cell(row=r, column=11).font = Font(name="Arial", size=8, color=BLUE)

row = 6
for i, (name, global_usd, sa_actual, trend, source, notes) in enumerate(industries):
    r = row + i
    ws.cell(row=r, column=1, value=name).font = bold_font

    # Col B: Global USD (millions)
    cell_b = ws.cell(row=r, column=2, value=global_usd)
    cell_b.font = blue_font
    cell_b.number_format = '$#,##0.00"M"'

    # Col C: SA ZAR (millions) — actual or formula-derived
    if sa_actual is not None:
        cell_c = ws.cell(row=r, column=3, value=sa_actual)
        cell_c.font = bold_font
        cell_c.fill = PatternFill("solid", fgColor=GREEN_BG)
    else:
        cell_c = ws.cell(row=r, column=3)
        cell_c.value = f"=B{r}*$K$4*$K$5"
        cell_c.font = Font(name="Arial", size=10, color=BLACK)
    cell_c.number_format = 'R#,##0.0"M"'

    # Col D: Cost per record = (SA breach cost in millions × 1000000) / avg records
    cell_d = ws.cell(row=r, column=4)
    cell_d.value = f"=C{r}*1000000/$K$3"
    cell_d.font = body_font
    cell_d.number_format = 'R#,##0'

    # Col E: Multiplier = SA breach cost / SA average
    cell_e = ws.cell(row=r, column=5)
    cell_e.value = f"=C{r}/$K$1"
    cell_e.font = body_font
    cell_e.number_format = '0.00"x"'

    # Col F: Trend
    cell_f = ws.cell(row=r, column=6, value=trend)
    cell_f.font = body_font
    if trend == "Up":
        cell_f.fill = PatternFill("solid", fgColor=RED_BG)
    else:
        cell_f.fill = PatternFill("solid", fgColor=GREEN_BG)

    # Col G: Source
    ws.cell(row=r, column=7, value=source).font = Font(name="Arial", size=9, color="64748B")

    # Col H: Notes
    ws.cell(row=r, column=8, value=notes).font = Font(name="Arial", size=9, color="64748B")

    # Apply borders
    for c in range(1, 9):
        ws.cell(row=r, column=c).border = border
        ws.cell(row=r, column=c).alignment = Alignment(vertical="center", wrap_text=(c == 8))

    # Alternate row shading
    if i % 2 == 1:
        for c in range(1, 9):
            if ws.cell(row=r, column=c).fill == PatternFill():
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=GREY_BG)

    ws.row_dimensions[r].height = 28

# Summary row
last_row = row + len(industries)
ws.merge_cells(f"A{last_row}:A{last_row}")
ws.cell(row=last_row, column=1, value="CROSS-INDUSTRY AVERAGE").font = Font(name="Arial", bold=True, size=10, color=WHITE)
ws.cell(row=last_row, column=1).fill = PatternFill("solid", fgColor=BLUE)
ws.cell(row=last_row, column=2, value=GLOBAL_AVG_USD_M).font = Font(name="Arial", bold=True, size=10, color=WHITE)
ws.cell(row=last_row, column=2).fill = PatternFill("solid", fgColor=BLUE)
ws.cell(row=last_row, column=2).number_format = '$#,##0.00"M"'
ws.cell(row=last_row, column=3, value=SA_AVG_ZAR_M).font = Font(name="Arial", bold=True, size=10, color=WHITE)
ws.cell(row=last_row, column=3).fill = PatternFill("solid", fgColor=BLUE)
ws.cell(row=last_row, column=3).number_format = 'R#,##0.0"M"'
cell_avg_d = ws.cell(row=last_row, column=4)
cell_avg_d.value = f"=C{last_row}*1000000/$K$3"
cell_avg_d.font = Font(name="Arial", bold=True, size=10, color=WHITE)
cell_avg_d.fill = PatternFill("solid", fgColor=BLUE)
cell_avg_d.number_format = 'R#,##0'
ws.cell(row=last_row, column=5, value=1.0).font = Font(name="Arial", bold=True, size=10, color=WHITE)
ws.cell(row=last_row, column=5).fill = PatternFill("solid", fgColor=BLUE)
ws.cell(row=last_row, column=5).number_format = '0.00"x"'
for c in range(6, 9):
    ws.cell(row=last_row, column=c).fill = PatternFill("solid", fgColor=BLUE)
for c in range(1, 9):
    ws.cell(row=last_row, column=c).border = border
    ws.cell(row=last_row, column=c).alignment = Alignment(vertical="center")

# Legend
lr = last_row + 2
ws.cell(row=lr, column=1, value="Legend:").font = Font(name="Arial", bold=True, size=9)
ws.cell(row=lr+1, column=1, value="Green cells = IBM SA direct data (actual reported figures)")
ws.cell(row=lr+1, column=1).font = Font(name="Arial", size=8, color="64748B")
ws.cell(row=lr+1, column=2).fill = PatternFill("solid", fgColor=GREEN_BG)
ws.cell(row=lr+2, column=1, value="White cells = Derived from global data using SA adjustment ratio (0.602)")
ws.cell(row=lr+2, column=1).font = Font(name="Arial", size=8, color="64748B")
ws.cell(row=lr+3, column=1, value="Blue text = Hardcoded inputs (IBM report data)")
ws.cell(row=lr+3, column=1).font = Font(name="Arial", size=8, color=BLUE)
ws.cell(row=lr+4, column=1, value="SA Adjustment Ratio = SA Avg (R44.1M) / (Global Avg $4.44M x R16.50) = 0.602")
ws.cell(row=lr+4, column=1).font = Font(name="Arial", size=8, color="64748B")

# Freeze panes
ws.freeze_panes = "A6"

# Print settings
ws.sheet_properties.pageSetUpPr.fitToPage = True

output = "SA_Data_Breach_Costs_by_Industry_2025.xlsx"
wb.save(output)
print(f"Spreadsheet saved: {output}")
